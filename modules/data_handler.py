"""
数据存储模块
负责保存详细的统计数据和结果导出

支持导出格式：TXT、Excel
"""
import json
import os
import shutil
from datetime import datetime
from modules.barcode_searcher import BarcodeSearcher


class DataHandler:
    """数据处理器"""

    RESULTS_FILE = "results.json"

    def __init__(self):
        self.current_session = {}

    def start_session(self, folder_path, shift='', production_line=''):
        """
        开始新的统计会话

        Args:
            folder_path: 图片文件夹路径
            shift: 统计班次（可选）
            production_line: 产线名称（可选）
        """
        self.current_session = {
            'folder_path': folder_path,
            'start_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'shift': shift,              # 新增：统计班次
            'production_line': production_line,  # 新增：产线名称
            'summary': {},
            'type_statistics': {},
            'details': []
        }

    def record_result(self, image_name, result, types=None, reason=None):
        """
        记录单张图片的统计结果

        Args:
            image_name: 图片文件名
            result: 判定结果 ('misjudgment' 或 'detection')
            types: 误判类型列表（仅当result为'misjudgment'时有值）
            reason: 误判原因（仅当result为'misjudgment'时有值）
        """
        record = {
            'image_name': image_name,
            'result': result,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        if result == 'misjudgment':
            if types:
                record['types'] = types
            if reason and reason.strip():  # 只在有值时添加
                record['reason'] = reason.strip()
        elif result == 'detection':
            # 检出时若有类型，记录类型用于统计回退与导出
            if types:
                record['types'] = types

        self.current_session['details'].append(record)

    def update_result(self, image_name, result, types=None, reason=None):
        """
        更新单张图片的标注结果（先删除旧记录再添加新记录）

        Args:
            image_name: 图片文件名
            result: 判定结果 ('misjudgment' 或 'detection')
            types: 误判类型列表（仅当result为'misjudgment'时有值）
            reason: 误判原因（仅当result为'misjudgment'时有值）

        Returns:
            dict: 返回旧记录（如果存在），否则返回None
        """
        # 先查找并删除旧记录
        old_record = self.remove_result(image_name)

        # 添加新记录
        self.record_result(image_name, result, types, reason)

        return old_record

    def remove_result(self, image_name):
        """
        删除指定图片的标注记录

        Args:
            image_name: 图片文件名

        Returns:
            dict: 返回被删除的旧记录（如果存在），否则返回None
        """
        details = self.current_session.get('details', [])
        for i, record in enumerate(details):
            if record['image_name'] == image_name:
                # 返回并删除旧记录
                return details.pop(i)
        return None

    def save_session(self, summary, type_statistics):
        """
        保存整个会话数据到JSON文件

        Args:
            summary: 统计摘要数据
            type_statistics: 各类型统计数据
        """
        self.current_session['end_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.current_session['summary'] = summary
        self.current_session['type_statistics'] = type_statistics

        try:
            with open(self.RESULTS_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.current_session, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存结果文件时发生错误: {e}")
            return False

    def load_session(self):
        """
        加载之前的会话数据

        Returns:
            dict: 会话数据，如果文件不存在返回None
        """
        if os.path.exists(self.RESULTS_FILE):
            try:
                with open(self.RESULTS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"加载结果文件时发生错误: {e}")
                return None
        return None

    def export_to_txt(self, output_file=None):
        """
        导出统计报告为TXT格式

        Args:
            output_file: 输出文件路径，如果为None则使用默认文件名

        Returns:
            bool: 导出成功返回True
        """
        if output_file is None:
            output_file = f"统计报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                # 写入基本信息
                f.write("=" * 50 + "\n")
                f.write("误判统计报告\n")
                f.write("=" * 50 + "\n\n")

                f.write(f"文件夹路径: {self.current_session.get('folder_path', 'N/A')}\n")
                f.write(f"开始时间: {self.current_session.get('start_time', 'N/A')}\n")
                f.write(f"结束时间: {self.current_session.get('end_time', 'N/A')}\n\n")

                # 写入统计摘要
                summary = self.current_session.get('summary', {})
                f.write("=" * 50 + "\n")
                f.write("统计摘要\n")
                f.write("=" * 50 + "\n")
                f.write(f"总图片数: {summary.get('total', 0)}\n")
                f.write(f"误判数量: {summary.get('misjudgment', 0)}\n")
                f.write(f"检出数量: {summary.get('detection', 0)}\n")
                f.write(f"误判率: {summary.get('misjudgment_rate', 0)}%\n")
                f.write(f"检出率: {summary.get('detection_rate', 0)}%\n\n")

                # 写入各类型统计
                type_stats = self.current_session.get('type_statistics', {})
                if type_stats:
                    f.write("=" * 50 + "\n")
                    f.write("各类型误判率\n")
                    f.write("=" * 50 + "\n")
                    for type_name, rate in type_stats.items():
                        f.write(f"{type_name}: {rate}%\n")
                    f.write("\n")

                # 写入详细记录
                details = self.current_session.get('details', [])
                f.write("=" * 50 + "\n")
                f.write(f"详细记录 (共{len(details)}条)\n")
                f.write("=" * 50 + "\n")
                for detail in details:
                    f.write(f"[{detail['timestamp']}] {detail['image_name']}: ")
                    f.write(f"误判" if detail['result'] == 'misjudgment' else "检出")
                    if detail['result'] == 'misjudgment' and 'types' in detail:
                        f.write(f" ({', '.join(detail['types'])})")
                    f.write("\n")

            return True
        except Exception as e:
            print(f"导出TXT报告时发生错误: {e}")
            return False

    def export_to_excel(self, output_file=None, total_images_count=0):
        """
        导出统计报告为Excel格式（增强版：包含班次/产线信息和双饼图）

        Args:
            output_file: 输出文件路径，如果为None则使用默认文件名
            total_images_count: 文件夹内所有图片数量（NG图片数）

        Returns:
            bool: 导出成功返回True
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.chart import PieChart, Reference
            from openpyxl.chart.label import DataLabelList
        except ImportError:
            print("未安装openpyxl库，请运行: pip install openpyxl")
            return False

        if output_file is None:
            output_file = f"统计报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "误判统计报告"

            # ========== 定义样式 ==========
            title_font = Font(size=16, bold=True, color='FFFFFF')
            section_font = Font(size=12, bold=True, color='4472C4')
            header_font = Font(size=11, bold=True, color='FFFFFF')
            content_font = Font(size=10)
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

            current_row = 1

            # ========== 1. 标题 ==========
            ws.merge_cells(f'A{current_row}:E{current_row}')
            title_cell = ws.cell(row=current_row, column=1, value='误判统计报告')
            title_cell.font = title_font
            title_cell.alignment = center_align
            title_cell.fill = title_fill
            current_row += 2

            # ========== 2. 基本信息 ==========
            ws.cell(row=current_row, column=1, value='基本信息').font = section_font
            current_row += 1

            # 准备基本信息数据
            shift = self.current_session.get('shift', '')
            line = self.current_session.get('production_line', '')
            info_data = [
                ('统计班次', shift if shift else '未设置'),
                ('产线名称', line if line else '未设置'),
                ('开始时间', self.current_session.get('start_time', 'N/A')),
                ('结束时间', self.current_session.get('end_time', 'N/A')),
                ('图片文件夹', self.current_session.get('folder_path', 'N/A'))
            ]

            for label, value in info_data:
                cell_label = ws.cell(row=current_row, column=1, value=label)
                cell_value = ws.cell(row=current_row, column=2, value=value)
                for cell in [cell_label, cell_value]:
                    cell.font = content_font
                    cell.alignment = left_align
                current_row += 1

            current_row += 1

            # ========== 3. 统计摘要 ==========
            ws.cell(row=current_row, column=1, value='统计摘要').font = section_font
            current_row += 1

            # 表头
            summary_headers = ['总产能', 'NG图片数', '误判图片总数', '总误判率(%)']
            for col, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            # 数据
            current_row += 1
            summary = self.current_session.get('summary', {})
            total_capacity = summary.get('total_capacity', 0)
            misjudgment_count = summary.get('misjudgment', 0)
            misjudgment_rate = summary.get('misjudgment_rate', 0)

            summary_values = [
                total_capacity,
                total_images_count,  # NG图片数：文件夹内所有图片数量
                misjudgment_count,
                misjudgment_rate
            ]

            for col, value in enumerate(summary_values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = content_font
                cell.alignment = center_align
                cell.border = thin_border

            current_row += 2

            # ========== 4. 各类型误判统计 ==========
            ws.cell(row=current_row, column=1, value='各类型误判统计').font = section_font
            current_row += 1

            # 表头
            type_headers = ['误判类型', '数量', '占比(%)', '误判率(%)']
            for col, header in enumerate(type_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            # 数据行
            current_row += 1
            type_stats = self.current_session.get('type_statistics', {})
            type_counts = summary.get('type_counts', {})
            detection_type_counts = summary.get('detection_type_counts', {})
            detection_count = summary.get('detection', 0)

            # 按数量降序排序
            sorted_types = sorted(type_counts.items(), key=lambda x: x[1], reverse=True)

            # 记录数据起始行（用于饼图）
            data_start_row = current_row

            for type_name, count in sorted_types:
                rate = type_stats.get(type_name, 0)
                # 计算占比（占总误判数的百分比）
                percentage = (count / misjudgment_count * 100) if misjudgment_count > 0 else 0

                cell1 = ws.cell(row=current_row, column=1, value=type_name)
                cell2 = ws.cell(row=current_row, column=2, value=count)
                cell3 = ws.cell(row=current_row, column=3, value=round(percentage, 2))
                cell4 = ws.cell(row=current_row, column=4, value=rate)

                for cell in [cell1, cell2, cell3, cell4]:
                    cell.font = content_font
                    cell.alignment = center_align
                    cell.border = thin_border

                current_row += 1

            data_end_row = current_row - 1

            # ========== 5. 各类型检出统计 ==========
            current_row += 1
            ws.cell(row=current_row, column=1, value='各类型检出统计').font = section_font
            current_row += 1

            # 表头
            detection_headers = ['检出类型', '数量', '占比(%)', '检出率(%)']
            for col, header in enumerate(detection_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            # 数据行
            current_row += 1
            sorted_detection_types = sorted(detection_type_counts.items(), key=lambda x: x[1], reverse=True)
            detection_data_start_row = current_row

            for type_name, count in sorted_detection_types:
                # 计算占比（占总检出数的百分比）
                percentage = (count / detection_count * 100) if detection_count > 0 else 0

                cell1 = ws.cell(row=current_row, column=1, value=type_name)
                cell2 = ws.cell(row=current_row, column=2, value=count)
                cell3 = ws.cell(row=current_row, column=3, value=round(percentage, 2))
                # 检出率列保留为空
                cell4 = ws.cell(row=current_row, column=4, value='')

                for cell in [cell1, cell2, cell3, cell4]:
                    cell.font = content_font
                    cell.alignment = center_align
                    cell.border = thin_border

                current_row += 1

            detection_data_end_row = current_row - 1

            # ========== 6. 饼图区域 ==========
            current_row += 1

            # 只在有误判数据时创建饼图
            if misjudgment_count > 0 and sorted_types:
                try:
                    # 饼图1：数量分布
                    pie1 = PieChart()
                    pie1.title = "误判类型数量分布"

                    labels = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
                    data1 = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)

                    pie1.add_data(data1, titles_from_data=False)
                    pie1.set_categories(labels)
                    pie1.width = 10
                    pie1.height = 8

                    # 设置数据标签
                    pie1.dataLabels = DataLabelList()
                    pie1.dataLabels.showVal = True
                    pie1.dataLabels.showPercent = True

                    # 将饼图固定在F2位置（避免挡住E列"误判原因"）
                    ws.add_chart(pie1, "F2")

                except Exception as e:
                    print(f"创建饼图时出现错误: {e}")

            # 只在有检出类型数据时创建饼图
            if detection_count > 0 and sorted_detection_types:
                try:
                    pie2 = PieChart()
                    pie2.title = "检出类型数量分布"

                    detection_labels = Reference(ws, min_col=1, min_row=detection_data_start_row, max_row=detection_data_end_row)
                    detection_data = Reference(ws, min_col=2, min_row=detection_data_start_row, max_row=detection_data_end_row)

                    pie2.add_data(detection_data, titles_from_data=False)
                    pie2.set_categories(detection_labels)
                    pie2.width = 10
                    pie2.height = 8

                    # 设置数据标签
                    pie2.dataLabels = DataLabelList()
                    pie2.dataLabels.showVal = True
                    pie2.dataLabels.showPercent = True

                    # 放在检出统计表下方的F列区域
                    ws.add_chart(pie2, f"F{current_row}")

                except Exception as e:
                    print(f"创建饼图时出现错误: {e}")

            # ========== 7. 详细记录 ==========
            details = self.current_session.get('details', [])
            ws.cell(row=current_row, column=1, value=f'详细记录 (共{len(details)}条)').font = section_font
            current_row += 1

            # 表头
            detail_headers = ['时间', '图片名称', '判定结果', '误判类型', '误判原因']
            for col, header in enumerate(detail_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border

            # 数据
            current_row += 1
            for detail in details:
                result_text = '误判' if detail['result'] == 'misjudgment' else '检出'
                types_text = ', '.join(detail.get('types', [])) if detail['result'] == 'misjudgment' else ''

                # 获取误判原因
                reason_text = detail.get('reason', '') if detail['result'] == 'misjudgment' else ''

                ws.cell(row=current_row, column=1, value=detail.get('timestamp', ''))
                ws.cell(row=current_row, column=2, value=detail.get('image_name', ''))
                ws.cell(row=current_row, column=3, value=result_text)
                ws.cell(row=current_row, column=4, value=types_text)
                ws.cell(row=current_row, column=5, value=reason_text)

                for col in range(1, 6):  # 修改为 1-5 列
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = content_font
                    cell.alignment = left_align
                    cell.border = thin_border

                current_row += 1

            # ========== 8. 自动调整列宽 ==========
            try:
                column_widths = {
                    'A': 20,  # 时间
                    'B': 30,  # 图片名称
                    'C': 12,  # 判定结果
                    'D': 40,  # 误判类型
                    'E': 30,  # 误判原因（新增）
                    'F': 15   # 预留
                }

                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
            except Exception as e:
                print(f"调整列宽时出现警告: {e}")

            wb.save(output_file)
            return True
        except Exception as e:
            print(f"导出Excel报告时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False

    def get_image_result(self, image_name):
        """
        获取指定图片的标注结果

        Args:
            image_name: 图片文件名

        Returns:
            dict: 如果找到返回 {'result': 'detection'/'misjudgment', 'types': [...]}
                  如果未找到返回 None
        """
        details = self.current_session.get('details', [])
        for record in details:
            if record['image_name'] == image_name:
                return record
        return None

    def get_session_summary(self):
        """
        获取当前会话的摘要信息

        Returns:
            dict: 会话摘要
        """
        return {
            'folder_path': self.current_session.get('folder_path', ''),
            'start_time': self.current_session.get('start_time', ''),
            'recorded_count': len(self.current_session.get('details', []))
        }

    def export_images(self, output_folder, folder_path, progress_callback=None):
        """
        导出误判图片及其同条码图片
        
        Args:
            output_folder: 导出目标文件夹
            folder_path: 原始图片文件夹路径（用于搜索）
            progress_callback: 进度回调函数，参数为 (current, total, message)
            
        Returns:
            tuple: (success: bool, message: str, stats: dict)
                   stats 包含 {'total_barcodes': int, 'total_images': int, 'barcode_details': dict}
        """
        try:
            # 1. 获取所有误判图片列表
            details = self.current_session.get('details', [])
            misjudgment_images = [
                detail['image_name'] 
                for detail in details 
                if detail['result'] == 'misjudgment'
            ]
            
            if not misjudgment_images:
                return (False, "没有误判图片可导出", {})
            
            if progress_callback:
                progress_callback(0, 100, "正在提取条码...")
            
            # 2. 提取所有误判图片的条码（只处理文件名中包含"NG"的图片）
            misjudgment_barcodes = set()
            # 记录每个条码对应的误判图片列表（用于生成清单）
            barcode_to_misjudgment = {}
            
            for img_name in misjudgment_images:
                # 只处理文件名中包含"NG"的图片
                if not BarcodeSearcher.is_ng_image(img_name):
                    continue
                    
                barcode = BarcodeSearcher.extract_barcode(img_name)
                misjudgment_barcodes.add(barcode)
                if barcode not in barcode_to_misjudgment:
                    barcode_to_misjudgment[barcode] = []
                barcode_to_misjudgment[barcode].append(img_name)
            
            if progress_callback:
                progress_callback(10, 100, f"已提取 {len(misjudgment_barcodes)} 个条码，正在搜索同条码图片...")
            
            # 3. 在搜索根目录（向上2级）搜索所有同条码且包含"NG"的图片
            # 例如：选择 E:\Image\2026-01-28-白班\检测拼接总图\NG 时，
            # 会在 E:\Image\2026-01-28-白班\ 目录下搜索所有子文件夹中
            # 文件名同时包含条码和"NG"的图片
            search_root = BarcodeSearcher.get_search_root(folder_path)
            barcode_images = BarcodeSearcher.search_images_by_barcodes(
                search_root, 
                misjudgment_barcodes,
                ng_only=True  # 只搜索包含"NG"的图片
            )
            
            if progress_callback:
                progress_callback(30, 100, "搜索完成，正在准备导出...")
            
            # 4. 按条码创建文件夹并复制图片
            total_images = 0
            barcode_details = {}  # 用于生成清单
            
            barcode_count = len(misjudgment_barcodes)
            for idx, barcode in enumerate(misjudgment_barcodes):
                # 创建条码文件夹
                barcode_folder = os.path.join(output_folder, barcode)
                os.makedirs(barcode_folder, exist_ok=True)
                
                # 获取该条码的所有图片
                images = barcode_images.get(barcode, [])
                
                # 记录该条码的详细信息
                barcode_details[barcode] = {
                    'total_count': len(images),
                    'misjudgment_images': barcode_to_misjudgment.get(barcode, []),
                    'all_images': []
                }
                
                # 复制图片
                for img_path in images:
                    img_name = os.path.basename(img_path)
                    dest_path = os.path.join(barcode_folder, img_name)
                    
                    try:
                        shutil.copy2(img_path, dest_path)
                        barcode_details[barcode]['all_images'].append(img_name)
                        total_images += 1
                    except Exception as e:
                        print(f"复制文件失败: {img_path} -> {dest_path}, 错误: {e}")
                
                # 更新进度
                if progress_callback:
                    progress = 30 + int((idx + 1) / barcode_count * 60)
                    progress_callback(
                        progress, 
                        100, 
                        f"正在导出条码 {idx + 1}/{barcode_count}..."
                    )
            
            if progress_callback:
                progress_callback(90, 100, "正在生成导出清单...")
            
            # 5. 生成详细清单文件
            summary_file = os.path.join(output_folder, "导出清单.txt")
            self._generate_export_summary(
                summary_file, 
                barcode_details,
                output_folder
            )
            
            if progress_callback:
                progress_callback(100, 100, "导出完成！")
            
            # 返回统计信息
            stats = {
                'total_barcodes': len(misjudgment_barcodes),
                'total_images': total_images,
                'barcode_details': barcode_details
            }
            
            return (True, f"成功导出 {len(misjudgment_barcodes)} 个条码组，共 {total_images} 张图片", stats)
            
        except Exception as e:
            import traceback
            error_msg = f"导出图片时发生错误: {e}\n{traceback.format_exc()}"
            print(error_msg)
            return (False, str(e), {})
    
    def _generate_export_summary(self, summary_file, barcode_details, output_folder):
        """
        生成详细的导出清单文件
        
        Args:
            summary_file: 清单文件路径
            barcode_details: 条码详细信息字典
            output_folder: 导出文件夹路径
        """
        try:
            with open(summary_file, 'w', encoding='utf-8') as f:
                # 标题
                f.write("=" * 60 + "\n")
                f.write("误判图片导出清单\n")
                f.write("=" * 60 + "\n\n")
                
                # 基本信息
                f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"导出路径: {output_folder}\n\n")
                
                # 统计汇总
                total_barcodes = len(barcode_details)
                total_images = sum(detail['total_count'] for detail in barcode_details.values())
                total_misjudgments = sum(len(detail['misjudgment_images']) for detail in barcode_details.values())
                
                f.write(f"共导出 {total_barcodes} 个条码组，总计 {total_images} 张图片\n")
                f.write(f"其中误判图片 {total_misjudgments} 张\n\n")
                
                f.write("=" * 60 + "\n\n")
                
                # 按条码详细列出
                for barcode, detail in sorted(barcode_details.items()):
                    f.write(f"条码: {barcode}\n")
                    f.write(f"图片数量: {detail['total_count']}\n")
                    f.write(f"子文件夹: {barcode}\\\n\n")
                    
                    # 创建误判图片集合，方便标记
                    misjudgment_set = set(detail['misjudgment_images'])
                    
                    # 列出所有图片，标记误判图片
                    for img_name in sorted(detail['all_images']):
                        if img_name in misjudgment_set:
                            f.write(f"  [误判] {img_name}\n")
                        else:
                            f.write(f"         {img_name}\n")
                    
                    f.write("\n" + "-" * 60 + "\n\n")
                
                f.write("=" * 60 + "\n")
                f.write("导出完成\n")
                
        except Exception as e:
            print(f"生成导出清单时发生错误: {e}")
